"""Task adapters that wrap existing Dunhill scripts without duplicating them."""

from __future__ import annotations

import logging
import os
import re
import signal
import subprocess
import sys
import time
from datetime import date
from pathlib import Path
from typing import Iterable

from scripts.workflow_dag import TaskResult


QUICKBI_PREFIXES = {
    "tm_order": "BI_tm_t01_trade_order_line",
    "tm_refund_success": "BI_tm_trade_refund_info_allsuc_filter",
    "tm_refund_pending": "BI_tm_trade_refund_info_paydate_filter",
    "dtc_order": "BI_dtc_t01_trade_order_line",
    "dtc_refund": "BI_dtc_t01_trade_refund_info_allsuc_filter",
}

DATA_IMPORT_DIR = Path("/Users/novel/Projects/data-import")
DATA_IMPORT_PYTHON = DATA_IMPORT_DIR / ".venv" / "bin" / "python"

JYCM_SAVE_NAMES = {
    "445603": "dunhill_shop_d_recent_30d_",
    "225266": "dunhill_traffic_d_recent_1d_",
    "446524": "dunhill_client_d_recent_30d_",
    "445350": "dunhill_product_d_recent_1d_",
    "458866": "dunhill_product_traffic_d_recent_1d_",
    # 周度清单（周日/周一）：30 天窗口报表，命名对齐 config/missions.py weekly_jycm_rpts
    "453129": "dunhill_product_d_recent_30d_",
    "224051": "dunhill_product_traffic_d_recent_30d_",
    "226600": "dunhill_traffic_d_recent_30d_",
}

QUICKBI_UPLOAD_TARGETS = {
    "tm_order": "dunhill_BI订单源",
    "tm_refund_success": "dunhill_TM退款源_hive",
    "tm_refund_pending": "dunhill_TM退款源_hive",
    "dtc_order": "dunhill_DTC订单源_hive",
    "dtc_refund": "dunhill_DTC退款源_hive",
}

# ponytail: reconcile v1 uses backup-file evidence except tm_order (real DB keys);
# add per-source SQL counts when a source is proven to drift.
BACKUP_TARGETS = {
    "445603": "dunhill_tm取数源_backup",
    "225266": "dunhill_tm流量源_new",
    "446524": "dunhill_tm客户源",
    "445350": "dunhill_tm商品源",
    "458866": "dunhill_tm商品流量源",
    # 周度清单（周日/周一）：30d 与 1d 报表入同一张目标表，REPLACE 按日期覆盖
    "453129": "dunhill_tm商品源",
    "224051": "dunhill_tm商品流量源",
    "226600": "dunhill_tm流量源_new",
    "tm_order": "dunhill_BI订单源",
    "tm_refund_success": "dunhill_TM退款源_hive",
    "tm_refund_pending": "dunhill_TM退款源_hive",
    "dtc_order": "dunhill_DTC订单源_hive",
    "dtc_refund": "dunhill_DTC退款源_hive",
}

STEP2_TIMEOUTS = {
    "jycm": 900,
    "sycm": 1800,
    "quickbi": 1800,
    "upload": 1800,
    "reconcile": 900,
    "crawler": 1800,
    "alimama": 1800,
}


def _manage(*args: str) -> list[str]:
    return [str(DATA_IMPORT_PYTHON), "manage.py", *args]


def _uploader(*args: str) -> list[str]:
    return [str(DATA_IMPORT_PYTHON), "-m", "data_pipeline.processors.file_uploader", *args]


def _descendants(pid: int) -> list[int]:
    """列同 PG 所有 PID（覆盖已 reparent 到 init 的孙子）。

    start_new_session=True 后主进程独占 PG，所有沿父链继承 PG 的子孙（Chrome/Playwright）
    都在这个 PG 里。pgrep -P 只能追直接子进程，shell fork 后 exec 的孙子 ppid=1 拿不到；
    列同 PG 一步到位。
    """
    try:
        pgid = os.getpgid(pid)
        same_pg = subprocess.run(
            ["pgrep", "-g", str(pgid)], capture_output=True, text=True, timeout=3
        ).stdout.split()
    except Exception:
        return []
    return [int(p) for p in same_pg if p != str(pid)]


def _kill_process_tree(pid: int, sig: int) -> None:
    """递归列子孙、再按反向（叶子先杀）发信号。

    start_new_session=True 后子进程独立 PG，os.killpg 会被 PermissionError 挡住，
    shell `/bin/kill -- -pgid` 也被 BSD 语法拒（仅 Linux 支持）。
    最稳的方式：递归 pgrep -P 拿所有子孙 PID 一一杀。
    """
    # 先杀孙子重孙，再杀直接子进程，最后杀主进程
    descendants = _descendants(pid)
    for d_pid in reversed(descendants):
        try:
            os.kill(d_pid, sig)
        except ProcessLookupError:
            pass
    try:
        os.kill(pid, sig)
    except ProcessLookupError:
        pass


def run_command(command: list[str], cwd: Path, timeout: int, env: dict | None = None) -> subprocess.CompletedProcess[str]:
    """Run a subprocess with process-tree kill on timeout.

    之前 subprocess.run(timeout=...) 在 socket IO 卡死时无效：socket 在 OS 内核里阻塞，
    Python 永远不会触发 TimeoutExpired，主进程无限等。改 Popen + communicate(timeout)
    + `_kill_process_tree` 走 pgrep 反向追子孙，SIGTERM/SIGKILL 走整组。
    """
    proc = subprocess.Popen(
        command,
        cwd=str(cwd),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="replace",
        start_new_session=True,  # 独立 PG，子孙跟主进程在同一 PG 下能 kill
        env=env,
    )
    try:
        stdout, stderr = proc.communicate(timeout=timeout)
    except subprocess.TimeoutExpired:
        _kill_process_tree(proc.pid, signal.SIGTERM)
        try:
            stdout, stderr = proc.communicate(timeout=5)
        except subprocess.TimeoutExpired:
            _kill_process_tree(proc.pid, signal.SIGKILL)
            stdout, stderr = proc.communicate()
        raise
    return subprocess.CompletedProcess(
        command, proc.returncode, stdout or "", stderr or ""
    )


def step1_command(task_id: str, root_dir: Path, timeout: int = 420) -> list[str]:
    script = root_dir / "scripts" / "step1_collect_data.py"
    command = [sys.executable, "-u", str(script), "--timeout", str(timeout)]
    if task_id == "refund_export":
        return [*command, "--skip-live", "--skip-quickbi"]
    if task_id == "live_export":
        return [*command, "--skip-refund", "--skip-quickbi"]
    prefix = "quickbi_browser."
    if task_id.startswith(prefix) and task_id[len(prefix):] in QUICKBI_PREFIXES:
        return [
            *command,
            "--skip-refund",
            "--skip-live",
            "--quickbi-sources",
            task_id[len(prefix):],
        ]
    raise ValueError(f"Unsupported Step 1 task: {task_id}")


def _files_matching(download_dir: Path, task_id: str, before: set[Path]) -> list[Path]:
    if task_id == "refund_export":
        candidates: Iterable[Path] = download_dir.glob("*.xlsx")
        candidates = [path for path in candidates if re.match(r"^\d+_\d+_\d+\.xlsx$", path.name)]
    elif task_id == "live_export":
        candidates = [
            path
            for pattern in ("直播间大盘数据-*.xlsx", "直播分场次效果*.xlsx", "*订单明细*.xlsx", "*成交订单明细*.xlsx")
            for path in download_dir.glob(pattern)
        ]
    else:
        source = task_id.split(".", 1)[1]
        prefix = QUICKBI_PREFIXES[source]
        candidates = download_dir.glob(f"{prefix}*.xlsx")
    return sorted({path for path in candidates if path.is_file() and path not in before}, key=lambda path: path.stat().st_mtime)


def _error_type(output: str) -> str:
    lowered = output.lower()
    # 控件定位失败是前端改版/时序问题，与登录无关，须先于 auth 关键词判断，
    # 否则日志里偶现“登录”字样会被误标为 auth_required
    if "内不可点击" in output or "控件 " in output:
        return "ui_change"
    # 页面回读与期望不符（填日期被拒回退默认值等）：T-1 数据未释放时的典型形态，
    # 须先于 auth 判断——失败输出末尾的排查建议含"登录"字样（2026-08-31 误标事故）
    if "回读不一致" in output or "期望" in output and "回读" in output:
        return "data_not_ready"
    # 页面元素点击/等待超时（Playwright locator）。登录态失效也会表现成元素点不动，
    # 但 auth 判断交给 taobao_auth_check gate，这里如实归类为 ui 问题。
    # （2026-08-31 refund/live 的 TimeoutError 被完整输出前段的 auth 字样误标）
    if "TimeoutError" in output or "locator.click" in lowered or "locator resolved" in lowered:
        return "ui_timeout"
    if any(marker in lowered for marker in ("login", "auth", "登录", "cookie")):
        return "auth_required"
    if any(marker in lowered for marker in ("timeout", "timed out", "connection", "net::err")):
        return "transient_network"
    return "file_error"


def _row_count(path: Path) -> int | None:
    """读 Excel 数据行数；打不开返回 None（视为未知，不拦下载）。"""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True) if any(cell is not None for cell in _))
        wb.close()
        return rows
    except Exception:
        return None


QUICKBI_PREVIEW_LIMIT = 50


def _quarantine_truncated(path: Path) -> Path:
    """把 API 预览截断文件移到隔离目录：既躲开上传模板匹配，又作为"已判截断"的标记。"""
    quarantine_dir = path.parent / "truncated_previews"
    quarantine_dir.mkdir(exist_ok=True)
    target = quarantine_dir / path.name
    path.rename(target)
    logging.info(f"{path.name} 达到 {QUICKBI_PREVIEW_LIMIT} 行预览上限，移入 {quarantine_dir.name}/，转浏览器全量导出")
    return target


def _is_truncated_marker(download_dir: Path, filename: str) -> bool:
    return (download_dir / "truncated_previews" / filename).exists()


def run_step1_task(task_id: str, root_dir: Path, download_dir: Path, timeout: int = 420) -> TaskResult:
    # quickbi_browser.* 是 quickbi_api.* 的导出兜底；API 预览最多 50 行，达到上限说明被截断
    if task_id.startswith("quickbi_browser."):
        source = task_id.split(".", 1)[1]
        pattern = _daily_glob(QUICKBI_PREFIXES[source], date.today())
        backup_dir = _default_save_path() / BACKUP_TARGETS.get(source, "")
        existing = list(download_dir.glob(pattern))
        if not existing:
            existing = list(backup_dir.glob(pattern))
        if existing:
            rows = _row_count(existing[0])
            marker = _is_truncated_marker(download_dir, existing[0].name)
            if rows is None or rows < QUICKBI_PREVIEW_LIMIT or marker:
                # rows<50：API 数据完整；marker 存在：此文件已是浏览器全量导出（resume 场景）
                return TaskResult(task_id, "skipped", evidence={"skipped_by": f"quickbi_api.{source}", "found": existing[0].name, "rows": rows, "full_export": marker})
            # ≥50 行且无标记：API 截断文件，隔离后走浏览器全量导出
            _quarantine_truncated(existing[0])

    before = {path for path in download_dir.glob("*.xlsx") if path.is_file()}
    started = time.time()
    try:
        result = run_command(step1_command(task_id, root_dir, timeout), root_dir, timeout + 30)
    except subprocess.TimeoutExpired as error:
        return TaskResult.failed(task_id, "transient_network", retryable=True, evidence={"message": str(error)})

    output = f"{result.stdout}\n{result.stderr}"
    outputs = _files_matching(download_dir, task_id, before)
    if task_id == "live_export" and not outputs and result.returncode == 0:
        # export_live 当日已有所需文件时会跳过下载直接成功（DAG 重试场景），此时没有"新增"文件
        outputs = _files_matching(download_dir, task_id, set())
    if result.returncode != 0:
        return TaskResult.failed(task_id, _error_type(output), retryable=_error_type(output) == "transient_network", evidence={"exit_code": result.returncode, "log_tail": output[-1000:]})
    if outputs:
        # log_tail 留存：2026-08-30 live_export session 阶段静默空手但整体 success，
        # 内部 stdout 被丢弃导致无法回溯——成功路径也留最后 1500 字符
        return TaskResult.success(task_id, outputs=[str(path) for path in outputs], evidence={"new_files": len(outputs), "duration_seconds": round(time.time() - started, 2), "log_tail": output[-1500:]})
    if task_id.startswith("quickbi_browser."):
        # 脚本正常退出但没有新文件 = 页面查询为空（如 dtc_refund 近期无退款），正常业务结果
        return TaskResult.no_data(task_id, evidence={"rows": 0, "confirmed_by": "browser_export"})
    if re.search(r'"rows"\s*:\s*0|rows\s*[:=]\s*0|无数据|no data', output, re.IGNORECASE):
        return TaskResult.no_data(task_id, evidence={"rows": 0, "confirmed_by": "export_log"})
    return TaskResult.failed(task_id, "file_error", evidence={"message": "command succeeded but produced no new source file"})


def step2_command(task_id: str, run_day: date) -> tuple[list[str], int] | None:
    """Map a Step 2 DAG task to one existing data-import command."""
    today = run_day.strftime("%Y%m%d")
    if task_id.startswith("jycm_download."):
        report_id = task_id.split(".", 1)[1]
        save_name = JYCM_SAVE_NAMES.get(report_id)
        if not save_name:
            return None
        return _manage("jycm", report_id, save_name + today), STEP2_TIMEOUTS["jycm"]
    if task_id == "sycm_download":
        return _manage("sycm", "shop"), STEP2_TIMEOUTS["sycm"]
    if task_id.startswith("quickbi_api."):
        # 单源抓取：每个 task 只跑自己的源，5 个 task 真正并行
        # （manage.py quickbi --sources X 由 commit 1 在 data-import 仓库支持）
        source = task_id.split(".", 1)[1]
        return _manage("quickbi", "--sources", source), STEP2_TIMEOUTS["quickbi"]
    if task_id.startswith("targeted_upload."):
        source_task = task_id.split(".", 1)[1]
        if source_task == "refund_export":
            # 千牛后台退款导出 → 主退款源（mission 在 modules.dunhill.ali.order）
            return _uploader("--target", "dunhill_TM退款源"), STEP2_TIMEOUTS["upload"]
        if source_task == "live_export":
            # 直播三件套：大盘/场次/订单明细（mission 在 modules.dunhill.ali.livestream）
            return _manage("upload", "-m", "modules.dunhill.ali.livestream"), STEP2_TIMEOUTS["upload"]
        if source_task.startswith("jycm_download."):
            template = JYCM_SAVE_NAMES[source_task.split(".", 1)[1]]
            return _uploader("--template", template), STEP2_TIMEOUTS["upload"]
        if source_task == "sycm_download":
            # sycm_download 实际产出的是 dunhill_*_d_recent_* 文件，属 jycm 模块的任务
            return _manage("upload", "-m", "modules.dunhill.ali.jycm"), STEP2_TIMEOUTS["upload"]
        source = source_task.split(".", 1)[1]
        return _uploader("--target", QUICKBI_UPLOAD_TARGETS[source]), STEP2_TIMEOUTS["upload"]
    if task_id == "unmask_buyer_nicknames":
        return _manage("order", "update_mask"), STEP2_TIMEOUTS["reconcile"]
    if task_id == "fq_crawler":
        return _manage("order", "fq"), STEP2_TIMEOUTS["crawler"]
    if task_id == "nickname_crawler":
        return _manage("order", "nick"), STEP2_TIMEOUTS["crawler"]
    if task_id == "pfs_buyer_type":
        return _manage("order", "update_pfs"), STEP2_TIMEOUTS["crawler"]
    if task_id == "dtc_buyer_type":
        return _manage("order", "update_dtc"), STEP2_TIMEOUTS["crawler"]
    if task_id == "alimama_auth_refresh_if_needed":
        return _manage("alimama", "--refresh-auth"), 600
    if task_id == "alimama_import":
        return _manage("alimama"), STEP2_TIMEOUTS["alimama"]
    return None


def _daily_glob(prefix: str, run_day: date) -> str:
    """当日源文件名统一为 `{prefix}_{YYYYMMDD}*.xlsx`（crawler/下载均带下划线；JYCM_SAVE_NAMES 前缀本身以 _ 结尾）。

    中缀 `*`：QuickBI 浏览器兜底导出的文件名带 `_自助取数_HH_MM_SS` 段
    （2026-09-01 tm_refund_success verify 落空的事故），API 文件不带，两者都要匹配。
    """
    return f"{prefix.rstrip('_')}_*{run_day.strftime('%Y%m%d')}*.xlsx"


def _source_file_token(source_task: str, run_day: date) -> tuple[str, str] | None:
    """Return (glob pattern, backup target) for verify/reconcile tasks."""
    if source_task == "refund_export":
        # 千牛后台退款导出：文件名纯数字时间戳，无法从文件名判断日期——用" Downloads 中存在未消费文件"判定
        files = [p for p in _download_dir().glob("*.xlsx") if re.match(r"^\d+_\d+_\d+\.xlsx$", p.name)]
        return (next((p.name for p in files), ""), "dunhill_TM退款源") if files else (f"__none_{run_day}", "dunhill_TM退款源")
    if source_task == "live_export":
        return "直播间*.xlsx", "dunhill_直播取数源"
    if source_task.startswith("jycm_download."):
        report_id = source_task.split(".", 1)[1]
        save_name = JYCM_SAVE_NAMES.get(report_id)
        if not save_name:
            return None
        return _daily_glob(save_name, run_day), BACKUP_TARGETS.get(report_id, "")
    if source_task == "sycm_download":
        return _daily_glob("dunhill_*d_*", run_day), "dunhill_tm取数源_backup"
    if source_task.startswith("quickbi_api."):
        source = source_task.split(".", 1)[1]
        prefix = QUICKBI_PREFIXES[source]
        return _daily_glob(prefix, run_day), BACKUP_TARGETS[source]
    return None


def _download_dir() -> Path:
    # .env 注入的值可能是未展开的 ~/Downloads（相对路径，glob 永远为空），必须 expanduser
    return Path(__import__("os").environ.get("DOWNLOAD_DIR", str(Path.home() / "Downloads"))).expanduser()


def _default_save_path() -> Path:
    from scripts.step2_run_import import get_default_save_path

    return Path(get_default_save_path())


def run_step2_task(task_id: str, root_dir: Path, upstream_state: dict | None = None) -> TaskResult:
    run_day = date.today()
    started = time.time()

    if task_id == "taobao_auth_check":
        # 真实检测千牛 cookies：调用 data-import 项目的 check_taobao_cookie_health
        check = [str(DATA_IMPORT_PYTHON), "-c",
                 "import sys; sys.path.insert(0, 'src'); "
                 "from data_pipeline.crawler.base import check_taobao_cookie_health; "
                 "sys.exit(0 if check_taobao_cookie_health() else 1)"]
        try:
            result = subprocess.run(check, cwd=str(DATA_IMPORT_DIR), capture_output=True, text=True, timeout=60)
        except subprocess.TimeoutExpired as error:
            return TaskResult.failed(task_id, "transient_network", retryable=True, evidence={"message": str(error)})
        if result.returncode != 0:
            return TaskResult.failed(task_id, "auth_required", retryable=True,
                                     evidence={"message": "千牛 cookies 已过期，请运行 scripts/login/taobao_login.py",
                                               "log_tail": (result.stdout + result.stderr)[-500:]})
        return TaskResult.success(task_id, evidence={"gate": "check_taobao_cookie_health"})

    if task_id in ("alimama_auth_check", "database_reconcile"):
        # Gates enforced through dependent command exit codes / dependency status.
        return TaskResult.success(task_id, evidence={"gate": "delegated"})

    if task_id.startswith("source_verify."):
        source_task = task_id.split(".", 1)[1]
        token = _source_file_token(source_task, run_day)
        if not token:
            return TaskResult.failed(task_id, "migration_pending", evidence={"message": f"no source mapping for {source_task}"})
        pattern, backup_target = token
        files: list = []
        # 短重试：并发写/同步盘抖动会让 glob 瞬时落空（2026-08-22 事故），不该一次落空就拖死上传链
        for attempt in range(3):
            files = sorted(_download_dir().glob(pattern))
            if not files and backup_target:
                # 上传后文件被 move 到 backup；重试/续跑场景下 backup 里有当日产物同样算验证通过
                files = sorted((_default_save_path() / backup_target).glob(pattern if pattern != f"__none_{run_day}" else "*"))
            if files:
                return TaskResult.success(task_id, outputs=[str(path) for path in files], evidence={"files": [path.name for path in files]})
            time.sleep(1)
        # 上游 source 任务本身 no_data（如 dtc_refund 近期无 DTC 退款、浏览器导出空表）
        # 时，文件不存在是正常结果，verify 同样 no_data，不该拖死上传链
        upstream = (upstream_state or {}).get("tasks", {}).get(source_task, {})
        if upstream.get("status") == "no_data":
            return TaskResult.no_data(task_id, evidence={"message": f"上游 {source_task} 无数据，跳过 {pattern}"})
        return TaskResult.failed(task_id, "file_error", evidence={"message": f"no download matching {pattern}", "diagnostics": {
            "download_dir": str(_download_dir()),
            "download_dir_env": __import__("os").environ.get("DOWNLOAD_DIR"),
            "cwd": __import__("os").getcwd(),
            "dir_entries": len(list(_download_dir().iterdir())),
            "sample": sorted(p.name for p in _download_dir().iterdir())[:8],
        }})

    if task_id.startswith("database_reconcile."):
        source_task = task_id.split(".", 1)[1]
        token = _source_file_token(source_task, run_day)
        if not token:
            return TaskResult.failed(task_id, "migration_pending", evidence={"message": f"no source mapping for {source_task}"})
        pattern, backup_target = token
        if source_task == "refund_export":
            # 千牛退款源 verify 通过 = 当日文件已上传（verify 已兜底查过 backup）；
            # Downloads 无文件且 backup 无当日新增 = 当天没有退款导出任务，no_data 而非 failed
            backup_dir = _default_save_path() / backup_target
            fresh = [p.name for p in backup_dir.glob("*.xlsx") if p.stat().st_mtime >= time.mktime(run_day.timetuple())]
            if fresh:
                return TaskResult.success(task_id, evidence={"backup": fresh})
            return TaskResult.no_data(task_id, evidence={"message": "当日无千牛退款导出（Downloads 与 backup 均无新文件）"})
        if source_task == "quickbi_api.tm_order":
            from scripts.step2_run_import import verify_tm_order_database

            ok, detail = verify_tm_order_database(DATA_IMPORT_DIR, run_day)
            if not ok:
                return TaskResult.failed(task_id, "database_error", retryable=True, evidence={"message": detail})
            return TaskResult.success(task_id, evidence={"reconciled": detail})
        if source_task == "live_export":
            # 大盘+订单明细必须入库；场次实时源可选（可能当天无场次数据），缺失只记录不判失败
            required = ("dunhill_直播取数源", "dunhill_直播订单源")
            optional = "dunhill_直播场次实时源"
            today_ts = time.mktime(run_day.timetuple())
            found = {d: [p.name for p in (_default_save_path() / d).glob("*.xlsx") if p.stat().st_mtime >= today_ts] for d in (*required, optional)}
            if not found["dunhill_直播订单源"]:
                # 空表日：订单明细 0 数据行，upload 正常跳过、文件留在 Downloads——确认后等同已入库
                empty_orders = [p.name for p in _download_dir().glob("直播间成交订单明细*.xlsx") if (_row_count(p) or 0) == 0 and p.stat().st_mtime >= today_ts]
                if empty_orders:
                    found["dunhill_直播订单源"] = [f"{name}（空表，无直播成交）" for name in empty_orders]
            if all(found[d] for d in required):
                return TaskResult.success(task_id, evidence={"backup": found, "场次实时源缺失": not found[optional] or None})
            missing = [d for d in required if not found[d]]
            detail = f"backup missing for {missing}: 当日直播文件未入库（订单明细空表属正常，由 export_live 直接判 success）"
            return TaskResult.failed(task_id, "file_error", retryable=True, evidence={"message": detail, "found": found})
        if backup_target and list((_default_save_path() / backup_target).glob(pattern)):
            return TaskResult.success(task_id, evidence={"backup": backup_target, "pattern": pattern})
        # 上游 source 任务 no_data（如 dtc_refund 近期无退款，Downloads/backup 均无产物）
        # 时，backup 缺失是正常结果，reconcile 同样 no_data
        upstream = (upstream_state or {}).get("tasks", {}).get(source_task, {})
        if upstream.get("status") == "no_data":
            return TaskResult.no_data(task_id, evidence={"message": f"上游 {source_task} 无数据，无需对账 {pattern}"})
        return TaskResult.failed(task_id, "file_error", retryable=True, evidence={"message": f"backup missing: {backup_target or '?'}/{pattern}"})

    if task_id.startswith("quickbi_api."):
        source = task_id.split(".", 1)[1]
        pattern = _daily_glob(QUICKBI_PREFIXES[source], run_day)
        found = list(_download_dir().glob(pattern))
        if not found:
            found = list((_default_save_path() / BACKUP_TARGETS.get(source, "")).glob(pattern))
        if found and not _is_truncated_marker(_download_dir(), found[0].name):
            # 已有当日文件（爬虫一次抓 5 个源，其余任务复用）；backup 命中 = 已上传
            return TaskResult.success(task_id, outputs=[str(path) for path in found], evidence={"already_downloaded": pattern})
        mapping = _manage("quickbi"), STEP2_TIMEOUTS["quickbi"]
        result = _run_subprocess(task_id, mapping)
        if result is not None and result.status == "failed":
            return result
        found = list(_download_dir().glob(pattern))
        if not found:
            # 查询为空（如 dtc_refund 近期无退款）或 API 无产出：正常业务场景，让位 browser 兜底判定，
            # 不能 failed——failed 会把 verify/upload/reconcile 整条链 blocked
            return TaskResult.no_data(task_id, evidence={"reason": "api 未产出当日文件，转 browser 兜底", "pattern": pattern})
        rows = _row_count(found[0])
        if rows is not None and rows >= QUICKBI_PREVIEW_LIMIT:
            # 预览截断：隔离截断文件，no_data 让 quickbi_browser 走浏览器全量导出
            _quarantine_truncated(found[0])
            return TaskResult.no_data(task_id, evidence={"reason": f"达 {QUICKBI_PREVIEW_LIMIT} 行预览上限，转 browser 全量导出", "rows": rows})
        return TaskResult.success(task_id, outputs=[str(path) for path in found], evidence={"pattern": pattern, "rows": rows})

    mapping = step2_command(task_id, run_day)
    if mapping is None:
        return TaskResult.failed(task_id, "migration_pending", evidence={"message": "no adapter yet"})
    return _run_subprocess(task_id, mapping)


def _run_subprocess(task_id: str, mapping: tuple[list[str], int]) -> TaskResult:
    started = time.time()
    command, timeout = mapping
    env = {**os.environ, "PYTHONPATH": str(DATA_IMPORT_DIR / "src")}
    try:
        result = run_command(command, DATA_IMPORT_DIR, timeout, env=env)
    except subprocess.TimeoutExpired as error:
        return TaskResult.failed(task_id, "transient_network", retryable=True, evidence={"message": str(error), "pid_killed": True})
    output = f"{result.stdout}\n{result.stderr}"
    if result.returncode != 0:
        error_type = _error_type(output)
        return TaskResult.failed(task_id, error_type, retryable=error_type in ("transient_network", "auth_required"), evidence={"exit_code": result.returncode, "log_tail": output[-1000:]})
    return TaskResult.success(task_id, evidence={"command": command[-3:], "duration_seconds": round(time.time() - started, 2)})
