#!/usr/bin/env python3
"""Load and parse Dunhill weekly data Excel file.

Usage:
    python load_excel.py <excel_path> [--sheet <sheet_name>] [--rows <max_rows>] [--summary]

Output: JSON to stdout with parsed sheet data.
"""
import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl"}))
    sys.exit(1)


def detect_sheet_type(name, index, total_sheets):
    """Classify sheet as transformed, raw, or reference."""
    reference_keywords = ["channelmapping", "mapping", "lookup"]
    raw_keywords = ["数据源", "raw", "全店铺数据", "竞品数据"]
    name_lower = name.lower().strip()

    if any(kw in name_lower for kw in reference_keywords):
        return "reference"
    # Last 3 sheets are typically raw data
    if any(kw in name_lower for kw in raw_keywords):
        return "raw"
    if index >= total_sheets - 3:
        return "raw"
    return "transformed"


def load_sheet(wb, sheet_name, max_rows=None):
    """Load a single sheet and return headers + rows as list of dicts."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"headers": [], "data": [], "total_rows": 0, "total_cols": 0}

    # Find header row (first non-empty row)
    header_row = None
    for i, row in enumerate(rows):
        if any(cell is not None for cell in row):
            header_row = i
            break

    if header_row is None:
        return {"headers": [], "data": [], "total_rows": 0, "total_cols": 0}

    headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(rows[header_row])]
    data_rows = []
    limit = max_rows if max_rows else len(rows) - header_row - 1

    for row in rows[header_row + 1: header_row + 1 + limit]:
        if any(cell is not None for cell in row):
            row_dict = {}
            for j, val in enumerate(row):
                if j < len(headers):
                    row_dict[headers[j]] = val
            data_rows.append(row_dict)

    return {
        "headers": headers,
        "data": data_rows,
        "total_rows": len(data_rows),
        "total_cols": len(headers),
    }


def main():
    parser = argparse.ArgumentParser(description="Parse Dunhill Excel data source")
    parser.add_argument("excel_path", help="Path to Excel file")
    parser.add_argument("--sheet", help="Specific sheet name to load (default: all)")
    parser.add_argument("--rows", type=int, default=None, help="Max rows per sheet")
    parser.add_argument("--summary", action="store_true", help="Sheet summary only (no data)")
    args = parser.parse_args()

    path = Path(args.excel_path)
    if not path.exists():
        print(json.dumps({"error": f"File not found: {args.excel_path}"}))
        sys.exit(1)

    wb = openpyxl.load_workbook(str(path), data_only=True)
    all_sheets = wb.sheetnames

    if args.sheet:
        if args.sheet not in all_sheets:
            print(json.dumps({"error": f"Sheet '{args.sheet}' not found. Available: {all_sheets}"}))
            sys.exit(1)
        sheets_to_load = [args.sheet]
    else:
        sheets_to_load = all_sheets

    result = {"file": str(path), "sheets": {}}

    for name in sheets_to_load:
        idx = all_sheets.index(name)
        sheet_type = detect_sheet_type(name, idx, len(all_sheets))
        parsed = load_sheet(wb, name, max_rows=args.rows if not args.summary else 0)

        sheet_info = {
            "type": sheet_type,
            "total_rows": parsed["total_rows"],
            "total_cols": parsed["total_cols"],
            "headers": parsed["headers"],
        }

        if not args.summary and not args.rows:
            # For full load without row limit, only include row count (data too large for JSON)
            sheet_info["data_included"] = False
            sheet_info["note"] = "Use --rows N to include data rows"
        elif parsed["data"]:
            sheet_info["data"] = parsed["data"]
            sheet_info["data_included"] = True

        result["sheets"][name] = sheet_info

    wb.close()
    print(json.dumps(result, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
